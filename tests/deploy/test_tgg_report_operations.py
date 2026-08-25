from __future__ import annotations

import hashlib
import importlib.util
import json
import os
import sqlite3
import threading
from io import BytesIO
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

import yaml
import pytest
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
DEPLOY = ROOT / "deploy" / "tgg" / "christopher"
PLUGIN = DEPLOY / "plugins" / "report-operations" / "__init__.py"
SMOKE = DEPLOY / "scripts" / "run_isolated_smoke.py"


def _load_plugin():
    spec = importlib.util.spec_from_file_location("christopher_report_operations", PLUGIN)
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(module)
    return module


def _load_smoke():
    spec = importlib.util.spec_from_file_location("christopher_isolated_smoke", SMOKE)
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(module)
    return module


class _API(BaseHTTPRequestHandler):
    token = "fixture-token"
    workbook = b"fixture-xlsx-bytes"
    requests: list[tuple[str, dict]] = []

    def do_POST(self):
        assert self.headers["Authorization"] == f"Bearer {self.token}"
        assert self.headers["X-PS-Tenant"] == "tgg"
        payload = json.loads(self.rfile.read(int(self.headers.get("Content-Length", "0"))) or b"{}")
        type(self).requests.append((self.path, payload))
        digest = hashlib.sha256(self.workbook).hexdigest()
        port = self.server.server_port
        responses = {
            "/fetch-sources": {
                "fetch_id": "fetch-1",
                "sources": [{
                    "name": "master",
                    "hash": digest,
                    "bytes": len(self.workbook),
                    "fetched_at": "now",
                    "sheet_tabs": ["AMK", "HG", "PG", "SK"],
                    "ref": f"http://127.0.0.1:{port}/files/master.xlsx",
                    "file_name": "master.xlsx",
                }],
                "preview_rows": {"master": [{"row": 1}]},
            },
            "/preview-reconcile": {
                "run_id": "run-1",
                "delta": {
                    "new_cases": 1,
                    "canonical_updates": 2,
                    "source_selection_updates": 6,
                    "work_items_inserted": 0,
                },
                "structured": {
                    "casesInBuild": 6,
                    "casesEvaluated": 6,
                    "casesCreated": 1,
                    "casesCanonicalChanged": 2,
                    "casesSourceSelectionChanged": 6,
                    "workItemsInserted": 0,
                },
                "warnings": {"structured_identity": []},
            },
            "/apply-reconcile": {
                "applied": {
                    "new_cases": 1,
                    "canonical_updates": 2,
                    "source_selection_updates": 6,
                    "work_items_inserted": 0,
                },
                "backup": {"path": "/server/backup", "hash": "b", "verified": True},
                "audit_batch_id": "audit-1",
            },
            "/generate": {"run_id": "run-1", "verdict": "pass", "checks": {"ok": True}, "reports": [{"zone": zone, "ref": f"ref-{zone}", "hash": digest} for zone in ("AMK", "HG", "PG", "SK")]},
            "/get-reports": {"files": [{"zone": zone, "ref": f"http://127.0.0.1:{port}/files/{zone}.xlsx", "hash": digest, "file_name": f"{zone}.xlsx"} for zone in ("AMK", "HG", "PG", "SK")], "receipt": {"window": "auto"}},
            "/status": {"stage": "generated", "ok": True, "populations_touched": {"cases": 6}},
        }
        body = json.dumps(responses[self.path]).encode()
        self.send_response(200)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        if self.path.startswith("/status"):
            assert self.headers["Authorization"] == f"Bearer {self.token}"
            assert self.headers["X-PS-Tenant"] == "tgg"
            type(self).requests.append((self.path, {}))
            body = json.dumps({"stage": "generated", "ok": True, "populations_touched": {"cases": 6}}).encode()
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return
        assert self.path.startswith("/files/")
        assert self.headers["User-Agent"] == "Christopher-TGG/1.0"
        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Length", str(len(self.workbook)))
        self.end_headers()
        self.wfile.write(self.workbook)

    def log_message(self, *_args):
        return


def test_six_verbs_round_trip_and_download_four_documents(tmp_path, monkeypatch):
    server = ThreadingHTTPServer(("127.0.0.1", 0), _API)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        module = _load_plugin()
        operations = {
            name: {"method": "POST", "path": f"/{name}"}
            for name in (
                "fetch-sources", "preview-reconcile", "apply-reconcile",
                "generate", "get-reports", "status",
            )
        }
        operations["status"]["method"] = "GET"
        section = {
            "enabled": True,
            "base_url": f"http://127.0.0.1:{server.server_port}",
            "allowed_download_hosts": [f"127.0.0.1:{server.server_port}"],
            "allow_insecure_downloads": True,
            "download_root": str(tmp_path),
            "headers": {"X-PS-Tenant": "tgg"},
            "auth": {"token_env": "CHRISTOPHER_TGG_PS_SERVICE_TOKEN", "scheme": "Bearer"},
            "operations": operations,
        }
        monkeypatch.setattr(module, "_section", lambda: section)
        monkeypatch.setattr(module, "read_raw_config", lambda: {
            "python_sandbox": {
                "datasets": {"media": {"type": "path", "path": str(tmp_path)}}
            }
        })
        monkeypatch.setenv("CHRISTOPHER_TGG_PS_SERVICE_TOKEN", _API.token)

        results = [
            module._fetch_sources({"cycle": "weekly"}),
            module._preview_reconcile({"fetch_id": "fetch-1"}),
            module._apply_reconcile({"run_id": "run-1"}),
            module._generate({"cycle": "weekly", "window": "auto"}),
            module._get_reports({"run_id": "run-1"}),
            module._status({"run_id": "run-1"}),
        ]
        assert all("error" not in json.loads(result) for result in results)
        sources = json.loads(results[0])["sources"]
        assert len(sources) == 1
        assert Path(sources[0]["local_path"]).read_bytes() == _API.workbook
        assert sources[0]["sandbox_path"] == "/inputs/media/fetch-1/master.xlsx"
        assert sources[0]["verified_hash"] == hashlib.sha256(_API.workbook).hexdigest()
        files = json.loads(results[4])["files"]
        assert len(files) == 4
        assert all(Path(item["local_path"]).read_bytes() == _API.workbook for item in files)
        assert [path for path, _ in _API.requests[-6:]] == [
            "/fetch-sources", "/preview-reconcile", "/apply-reconcile",
            "/generate", "/get-reports", "/status?run_id=run-1",
        ]
        assert _API.requests[-6:][0][1] == {"cycle": "weekly"}
        assert _API.requests[-6:][3][1] == {"cycle": "weekly", "window": "auto"}
    finally:
        server.shutdown()
        thread.join(timeout=2)


def test_corrupt_source_shape_stops_at_fetch(monkeypatch):
    module = _load_plugin()
    monkeypatch.setattr(module, "_request", lambda *_args, **_kwargs: {
        "fetch_id": "fetch-bad",
        "sources": [{"name": "half-save", "hash": "a"}],
        "preview_rows": [],
    })
    result = json.loads(module._fetch_sources({"cycle": "weekly"}))
    assert "error" in result
    assert "response missing" in result["error"]


def test_plugin_registers_exact_six_typed_tools():
    module = _load_plugin()
    registered = []

    class Context:
        def register_tool(self, **kwargs):
            registered.append(kwargs)

    module.register(Context())
    assert [item["name"] for item in registered] == [
        "report_fetch_sources",
        "report_preview_reconcile",
        "report_apply_reconcile",
        "report_generate",
        "report_get_reports",
        "report_status",
    ]
    assert all(item["toolset"] == "report-operations" for item in registered)
    assert all(item["schema"]["parameters"]["additionalProperties"] is False for item in registered)


def test_client_surface_config_and_schedule_are_disabled():
    config = yaml.safe_load((DEPLOY / "config.yaml").read_text())
    report = config["pa"]["report_operations"]
    assert report["enabled"] is True
    assert report["schedule"]["enabled"] is False
    assert all(operation["method"] == "POST" for operation in report["operations"].values())
    assert all(operation["path"].endswith("?tenant=tgg") for operation in report["operations"].values())
    assert config["plugins"]["enabled"] == ["report-operations"]
    constitution = yaml.safe_load((DEPLOY / "christopher_tgg_constitution.yaml").read_text())
    management = constitution["job_briefs"]["tgg_management"]
    ingest = constitution["job_briefs"]["tgg_ops_ingest"]
    assert "report-operations" in management["enabled_toolsets"]
    assert "report-operations" not in ingest["enabled_toolsets"]
    text = "\n".join(management["instructions"])
    assert '"run weekly report"' in text
    assert '"run monthly report"' in text
    assert '"retry report run <id>"' in text
    assert "STOP: do not preview" in text
    assert "For report runs on 3 August 2026 only" in text
    assert "do not call report_apply_reconcile" in text
    assert "expires after 3 August 2026" in text


def test_bootstrap_makes_new_report_plugin_directories_traversable():
    bootstrap = (DEPLOY / "scripts" / "bootstrap_runtime.sh").read_text()
    assert "install -d -m 0755 -o root -g root" in bootstrap
    assert '"$DEPLOY_ROOT/plugins/report-operations"' in bootstrap


def test_isolated_smoke_exports_fixture_token_before_plugin_discovery(
    monkeypatch, tmp_path
):
    script = DEPLOY / "scripts" / "run_isolated_smoke.py"
    spec = importlib.util.spec_from_file_location("christopher_report_smoke", script)
    smoke = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(smoke)

    source = tmp_path / "live.env"
    target = tmp_path / "fixture.env"
    source.write_text("OPENAI_API_KEY=fixture-openai\n", encoding="utf-8")
    monkeypatch.delenv("CHRISTOPHER_TGG_PS_SERVICE_TOKEN", raising=False)

    smoke._copy_test_env(source, target)

    assert os.environ["CHRISTOPHER_TGG_PS_SERVICE_TOKEN"] == "fixture-only"


def test_runtime_verifier_ignores_non_report_setup_requests():
    source = (DEPLOY / "scripts" / "verify_runtime.sh").read_text()
    assert source.count('if path.startswith("/api/operator/report-cycle/")') == 2
    assert 'clean["report_ops_request_paths"][:6]' not in source
    assert 'corrupt["report_ops_request_paths"] ==' not in source


def test_isolated_smoke_serves_downloadable_master_and_closure_sources():
    smoke = _load_smoke()
    smoke._OperatorStub.workbooks = {}
    smoke._OperatorStub.report_scenario = "clean"
    server = ThreadingHTTPServer(("127.0.0.1", 0), smoke._OperatorStub)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        import urllib.request

        with urllib.request.urlopen(
            f"http://127.0.0.1:{server.server_port}/api/operator/report-cycle/fetch-sources?tenant=tgg"
        ) as response:
            source = json.loads(response.read())
        master = next(item for item in source["sources"] if item["name"] == "master")
        assert master["sheet_tabs"] == ["AMK", "HG", "PG", "SK"]
        with urllib.request.urlopen(master["ref"]) as response:
            workbook = load_workbook(BytesIO(response.read()), read_only=True, data_only=True)
        assert workbook.sheetnames == ["AMK", "HG", "PG", "SK"]
        assert {item["name"] for item in source["sources"]} == {"master", "closure"}
    finally:
        server.shutdown()
        thread.join(timeout=2)


def test_isolated_smoke_workbooks_are_valid_and_inspectable():
    module = _load_smoke()
    payload = module._OperatorStub._workbook(["AMK", "Serangoon", "Bishan"])

    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)

    assert workbook.sheetnames == ["AMK", "Serangoon", "Bishan"]
    assert workbook["AMK"]["A2"].value == "AM/JOB/2608/0001"


def test_isolated_smoke_case_query_reads_the_copied_database_only(tmp_path):
    module = _load_smoke()
    database = tmp_path / "tgg.db"
    with sqlite3.connect(database) as connection:
        connection.execute("CREATE TABLE cases (job_no TEXT, due_date TEXT)")
        connection.executemany(
            "INSERT INTO cases VALUES (?, ?)",
            [("AM/JOB/2608/0001", "2026-08-25"), ("PG/JOB/2608/0002", "2026-08-26")],
        )
    module._OperatorStub.case_db = database
    module._OperatorStub.case_queries = []

    result = module._OperatorStub._run_case_query(
        "SELECT job_no, due_date FROM cases ORDER BY job_no"
    )

    assert result["data"]["rows"] == [
        {"job_no": "AM/JOB/2608/0001", "due_date": "2026-08-25"},
        {"job_no": "PG/JOB/2608/0002", "due_date": "2026-08-26"},
    ]
    assert module._OperatorStub.case_queries[0]["row_count"] == 2
    with pytest.raises(ValueError, match="read-only"):
        module._OperatorStub._run_case_query("DELETE FROM cases")
    with sqlite3.connect(database) as connection:
        assert connection.execute("SELECT count(*) FROM cases").fetchone()[0] == 2


def test_runtime_verifier_identifies_its_systems_read_check_to_the_edge():
    source = (DEPLOY / "scripts" / "verify_runtime.sh").read_text()
    assert source.count('"User-Agent": "Christopher-TGG/1.0"') == 2


def test_scheduled_runner_is_outbound_disabled_in_dry_run(monkeypatch, tmp_path):
    script = DEPLOY / "scripts" / "run_scheduled_report.py"
    source = script.read_text()
    assert "process_live_records" not in source
    assert "TGG_REPLY_BRIDGE_URL" not in source
    assert "if not args.dry_run:" in source
    assert "append_record(Path(args.source), record)" in source
    for timer in ("weekly", "monthly"):
        assert "OnCalendar=Mon " in (DEPLOY / "systemd" / f"christopher-tgg-report-{timer}.timer").read_text()
    bootstrap = (DEPLOY / "scripts" / "bootstrap_runtime.sh").read_text()
    assert 'if [[ "$schedule_enabled" == "true" ]]' in bootstrap
    assert "systemctl disable --now christopher-tgg-report-weekly.timer" in bootstrap

    spec = importlib.util.spec_from_file_location(
        "christopher_scheduled_report", DEPLOY / "scripts" / "run_scheduled_report.py"
    )
    runner = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(runner)
    record = runner.build_record("weekly", now=123)
    assert record["senderId"] == "system@internal"
    assert record["body"] == "[system] scheduled weekly report run"
    assert record["chatId"] == runner.MANAGEMENT_CHAT
    monkeypatch.setattr(runner, "append_record", lambda *_args: (_ for _ in ()).throw(AssertionError("dry-run appended")))
    monkeypatch.setattr("sys.argv", [str(script), "--cycle", "weekly", "--dry-run", "--timestamp", "123"])
    assert runner.main() == 0


def test_nightly_whatsapp_trigger_is_internal_idempotent_and_outbound_disabled(monkeypatch, tmp_path):
    script = DEPLOY / "scripts" / "run_nightly_whatsapp.py"
    spec = importlib.util.spec_from_file_location("christopher_nightly_whatsapp", script)
    runner = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(runner)

    record = runner.build_record("2026-08-17", now=123)
    assert record["chatId"] == "900000000000000001@g.us"
    assert record["senderId"] == "system@internal"
    assert record["body"] == "[system] process TGG WhatsApp batch for 2026-08-17"
    assert record["fromMe"] is False

    source = tmp_path / "events.jsonl"
    source.write_text("", encoding="utf-8")
    receipts = tmp_path / "receipts"
    monkeypatch.setattr(runner, "DEFAULT_SOURCE", str(source.resolve()))
    assert runner.append_once(source, receipts, record) is True
    assert runner.append_once(source, receipts, record) is False
    assert len(source.read_text(encoding="utf-8").splitlines()) == 1
    receipt = json.loads((receipts / "2026-08-17.json").read_text())
    assert receipt["external_outbound_sent"] == 0
    assert receipt["message_id"] == record["messageId"]

    timer = (DEPLOY / "systemd" / "christopher-tgg-nightly-whatsapp.timer").read_text()
    assert "00:15:00 Asia/Singapore" in timer
    service = (DEPLOY / "systemd" / "christopher-tgg-nightly-whatsapp.service").read_text()
    assert "User=pclaw" in service
    assert "SupplementaryGroups=tggcapture" in service
    assert (
        "ExecStartPre=+/usr/bin/chmod 0660 "
        "/var/lib/tgg-capture/whatsapp/capture/events.jsonl"
    ) in service


def test_shared_runtime_files_have_no_new_report_domain_adapter():
    # The adapter and judgment vocabulary stay below the per-client deploy root.
    changed_surface = [PLUGIN, DEPLOY / "patches" / "report-operations-management.snippet.yaml"]
    assert all(path.is_relative_to(DEPLOY) for path in changed_surface)
    assert not (ROOT / "tools" / "tgg_report_ops.py").exists()
