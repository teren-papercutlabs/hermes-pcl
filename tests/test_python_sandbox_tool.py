"""Contract and jailed integration tests for the offline Python sandbox."""

from __future__ import annotations

import hashlib
import errno
import json
import os
import platform
import shutil
import signal
import sqlite3
import subprocess
import sys
import threading
import time
import zipfile
from pathlib import Path
from unittest.mock import Mock

import pytest

import tools.python_sandbox_tool as sandbox


def _db(path: Path, count: int = 550) -> None:
    connection = sqlite3.connect(path)
    connection.execute(
        "create table records (record_no text primary key, state text not null)"
    )
    connection.executemany(
        "insert into records values (?, ?)",
        [(f"R{index:04d}", "open" if index % 2 else "closed") for index in range(count)],
    )
    connection.commit()
    connection.close()


def _csv(path: Path) -> None:
    lines = ["record_no,state"]
    # 550 known rows: exactly 20 have the opposite state.
    for index in range(550):
        actual = "open" if index % 2 else "closed"
        state = ("closed" if actual == "open" else "open") if index < 20 else actual
        lines.append(f"R{index:04d},{state}")
    # 50 unknown rows, for 600 total.
    for index in range(550, 600):
        lines.append(f"R{index:04d},open")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _config(db: Path, csv: Path, **limits) -> dict:
    return {
        "enabled": True,
        "datasets": {
            "records": {
                "type": "sqlite",
                "path": str(db),
                "description": "Generic record database.",
            },
            "sheet": {
                "type": "path",
                "path": str(csv),
                "description": "Generic comparison CSV.",
            },
        },
        "limits": limits,
    }


def test_limits_defaults_and_malformed_values():
    assert sandbox._limits({}) == sandbox.DEFAULTS
    parsed = sandbox._limits(
        {"limits": {"wall_seconds": "bad", "cpu_seconds": "7", "memory_mb": -2}}
    )
    assert parsed["wall_seconds"] == 120
    assert parsed["cpu_seconds"] == 7
    assert parsed["memory_mb"] == 1


def test_timeout_is_clamped_to_configured_bounds():
    config = {"limits": {"wall_seconds": 33, "max_wall_seconds": 45}}
    assert sandbox._wall_seconds(None, config) == 33
    assert sandbox._wall_seconds(1, config) == 5
    assert sandbox._wall_seconds(999, config) == 45
    assert sandbox._wall_seconds("bad", config) == 33


def test_workspace_scope_defaults_to_run_and_requires_explicit_session_id(monkeypatch):
    assert sandbox._workspace_mode({}) == "run"
    assert sandbox._workspace_mode({"workspace": "session"}) == "session"
    assert sandbox._workspace_mode({"workspace": "unexpected"}) == "run"
    monkeypatch.setattr(
        sandbox,
        "_load_config",
        lambda: {"enabled": True, "workspace": "session"},
    )
    monkeypatch.setattr(sandbox, "_probe", lambda force=False: (True, "ok"))
    response = json.loads(sandbox.python_sandbox("print(1)"))
    assert response["status"] == "error"
    assert "requires a session_id" in response["error"]


def test_handler_forwards_explicit_session_id(monkeypatch):
    monkeypatch.setattr(sandbox, "_probe", lambda force=False: (True, "ok"))
    captured = {}

    def fake_python_sandbox(*args, **kwargs):
        captured.update(kwargs)
        return json.dumps({"status": "success"})

    monkeypatch.setattr(sandbox, "python_sandbox", fake_python_sandbox)
    response = json.loads(
        sandbox._handle_python_sandbox({"code": "print(1)"}, session_id="chat-a")
    )
    assert response["status"] == "success"
    assert captured["session_id"] == "chat-a"


def test_unknown_dataset_lists_valid_names(tmp_path):
    db, csv = tmp_path / "records.db", tmp_path / "input.csv"
    _db(db)
    _csv(csv)
    mounts, error = sandbox._resolve_datasets(
        ["missing"], _config(db, csv), tmp_path / "inputs"
    )
    assert mounts == {}
    assert "unknown dataset" in error
    assert "records, sheet" in error


def test_sqlite_snapshot_is_consistent_and_source_untouched(tmp_path):
    source, csv = tmp_path / "records.db", tmp_path / "input.csv"
    _db(source)
    _csv(csv)
    before = (
        source.stat().st_mtime_ns,
        hashlib.sha256(source.read_bytes()).hexdigest(),
    )
    inputs = tmp_path / "inputs"
    mounts, error = sandbox._resolve_datasets(
        ["records"], _config(source, csv), inputs
    )
    assert error is None
    snapshot = mounts["records"]
    assert snapshot != source
    assert sqlite3.connect(snapshot).execute("select count(*) from records").fetchone() == (
        550,
    )
    after = (
        source.stat().st_mtime_ns,
        hashlib.sha256(source.read_bytes()).hexdigest(),
    )
    assert after == before


def test_snapshot_size_guard_does_not_leave_partial_file(tmp_path):
    source = tmp_path / "large.db"
    source.write_bytes(b"x" * 2048)
    destination = tmp_path / "snapshot.db"
    with pytest.raises(ValueError, match="max_snapshot_mb"):
        sandbox._snapshot_sqlite(source, destination, 0)
    assert not destination.exists()


def test_snapshot_is_consistent_while_wal_writer_holds_transaction(tmp_path):
    source = tmp_path / "records.db"
    _db(source, count=10)
    writer = sqlite3.connect(source)
    writer.execute("pragma journal_mode=wal")
    writer.execute("begin immediate")
    writer.execute("insert into records values ('UNCOMMITTED', 'open')")
    destination = tmp_path / "snapshot.db"
    sandbox._snapshot_sqlite(source, destination, 10)
    assert sqlite3.connect(destination).execute(
        "select count(*) from records"
    ).fetchone() == (10,)
    # The backup did not disturb or commit the live writer.
    assert writer.in_transaction is True
    assert writer.execute(
        "select count(*) from records where record_no='UNCOMMITTED'"
    ).fetchone() == (1,)
    writer.rollback()
    writer.close()


def test_path_symlink_escape_is_rejected(tmp_path):
    allowed = tmp_path / "allowed"
    allowed.mkdir()
    outside = tmp_path / "outside.csv"
    outside.write_text("secret", encoding="utf-8")
    link = allowed / "input.csv"
    link.symlink_to(outside)
    mounts, error = sandbox._resolve_datasets(
        ["input"],
        {"datasets": {"input": {"type": "path", "path": str(link)}}},
        tmp_path / "inputs",
    )
    assert mounts == {}
    assert "escapes allowed directory" in error


def test_init_script_mount_plan_has_one_write_surface(tmp_path):
    run = tmp_path / "run"
    (run / "inputs").mkdir(parents=True)
    (run / "work").mkdir()
    (run / "script.py").write_text("", encoding="utf-8")
    params = run / "inputs" / "params.json"
    params.write_text("{}", encoding="utf-8")
    file_input = run / "inputs" / "records.db"
    file_input.write_text("", encoding="utf-8")
    directory_input = tmp_path / "media"
    directory_input.mkdir()
    script = sandbox._generate_init_script(
        run,
        {"records": file_input, "media": directory_input},
        Path("/opt/runtime"),
        base_prefix=Path("/opt/python/3.13"),
        sqlite_datasets={"records"},
    )
    assert "--rbind" in script
    assert 'findmnt -Rrn -o TARGET "$dst"' in script
    assert 'remount,bind,ro "$target"' in script
    assert 'tmpfs "$JAIL/work"' in script
    assert 'mount -o remount,bind,ro "$JAIL/export"' in script
    assert 'ro_file ' in script and '"$JAIL/supervisor.py"' in script
    assert "exec /venv/bin/python -I /supervisor.py" in script
    assert "PR_SET_DUMPABLE = 4" in sandbox._SUPERVISOR_SOURCE
    assert '"remount,bind,rw", "/export"' in sandbox._SUPERVISOR_SOURCE
    assert sandbox._SUPERVISOR_SOURCE.count("remount,bind,rw") == 1
    assert "$HOME" not in script
    assert "HERMES_HOME" not in script
    assert '/usr/sbin/pivot_root "$JAIL"' in script
    assert 'mount -o remount,bind,ro "$JAIL"' in script
    assert '"/usr/bin/unshare"' in sandbox._SUPERVISOR_SOURCE
    assert "resource.setrlimit(resource.RLIMIT_NPROC" in sandbox._SUPERVISOR_SOURCE
    assert "cd /work" in script
    assert 'cp ' in script and '"$JAIL/work/records.db"' in script
    assert '"$JAIL/inputs/records"' not in script
    assert 'mkdir -p "$JAIL/opt/python/3.13"' in script
    assert 'ro_dir /opt/python/3.13 "$JAIL/opt/python/3.13"' in script


def test_system_base_prefix_needs_no_duplicate_mount(tmp_path):
    run = tmp_path / "run"
    (run / "inputs").mkdir(parents=True)
    (run / "work").mkdir()
    (run / "script.py").write_text("", encoding="utf-8")
    (run / "inputs" / "params.json").write_text("{}", encoding="utf-8")
    script = sandbox._generate_init_script(
        run, {}, Path("/opt/runtime"), base_prefix=Path("/usr/local")
    )
    assert 'ro_dir /usr/local "$JAIL/usr/local"' not in script


def test_child_env_scrubs_secrets_and_exposes_contract_paths(tmp_path):
    env = sandbox._build_env(
        {"records": tmp_path / "records.db", "media": tmp_path / "media"},
        {"PATH": "/bin", "SERVICE_TOKEN": "nope", "HERMES_TIMEZONE": "UTC"},
        sqlite_datasets={"records"},
    )
    assert "SERVICE_TOKEN" not in env
    assert json.loads(env["SANDBOX_INPUTS"]) == {
        "media": "/inputs/media",
        "records": "/work/records.db",
    }
    assert env["RESULT_PATH"] == "/work/result.json"
    assert env["TMPDIR"] == "/work"
    assert env["HOME"] == "/work"
    assert env["PATH"] == "/usr/sbin:/usr/bin:/sbin:/bin"


def test_stdout_head_tail_truncation_preserves_both_ends():
    text = "HEAD" + ("x" * 100) + "TAIL"
    value, truncated = sandbox._cap_head_tail(text, 20)
    assert truncated is True
    assert value.startswith("HEAD")
    assert value.endswith("TAIL")
    assert "[TRUNCATED:" in value
    assert "bytes /" in value
    assert "entries omitted]" in value


def test_concurrent_drain_handles_output_larger_than_pipe_buffer():
    size = 256 * 1024
    process = subprocess.Popen(
        [
            sys.executable,
            "-c",
            f"import sys; sys.stdout.buffer.write(b'H'+b'x'*{size}+b'T')",
        ],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    head, tail, total = [], [], [0]
    reader = threading.Thread(
        target=sandbox._drain_stream,
        args=(process.stdout, head, tail, total, 64, 64),
    )
    reader.start()
    assert process.wait(timeout=5) == 0
    reader.join(timeout=2)
    assert not reader.is_alive()
    output = sandbox._assemble_drain(head, tail, total[0], 128)
    assert total[0] == size + 2
    assert output.startswith("H")
    assert output.endswith("T")
    assert "chars omitted" in output


def test_result_cap_and_file_listing(tmp_path):
    work = tmp_path / "work"
    work.mkdir()
    (work / "result.json").write_text(
        json.dumps({"large": "x" * sandbox.RESULT_CAP}), encoding="utf-8"
    )
    (work / "details.csv").write_text("a\nb\n", encoding="utf-8")
    payload, error = sandbox._harvest(
        work, "summary", "", "success", sandbox.DEFAULTS
    )
    assert payload["status"] == "result_invalid"
    assert payload["truncated"]["result"] is True
    assert "cap 8KB" in error
    assert payload["files"] == [
        {"path": "work/details.csv", "bytes": 4, "lines": 2}
    ]


def test_xlsx_file_listing_includes_configured_client_url(tmp_path):
    work = tmp_path / "work"
    work.mkdir()
    (work / "weekly-report.xlsx").write_bytes(b"workbook")

    payload, error = sandbox._harvest(
        work,
        "summary",
        "",
        "success",
        sandbox.DEFAULTS,
        run_id="r_12ab34cd",
        artifact_url_base="https://portal.example/artifacts/",
    )

    assert error == ""
    assert payload["files"] == [
        {
            "path": "work/weekly-report.xlsx",
            "bytes": 8,
            "client_url": (
                "https://portal.example/artifacts/r_12ab34cd/weekly-report.xlsx"
            ),
            "lines": 1,
        }
    ]


def test_xlsx_file_listing_promotes_atomically_and_idempotently(tmp_path):
    work = tmp_path / "work"
    retained = tmp_path / "retained"
    work.mkdir()
    workbook = work / "weekly-report.xlsx"
    with zipfile.ZipFile(workbook, "w") as archive:
        archive.writestr("[Content_Types].xml", "<Types/>")

    kwargs = {
        "run_id": "r_12ab34cd",
        "artifact_url_base": "https://portal.example/artifacts",
        "media_retention": {
            "root": str(retained),
            "media_ref_prefix": "/media/tgg/hermes",
        },
    }
    first, error = sandbox._harvest(
        work, "summary", "", "success", sandbox.DEFAULTS, **kwargs
    )
    promoted = list(retained.glob("*.xlsx"))
    assert error == ""
    assert len(promoted) == 1
    assert promoted[0].read_bytes() == workbook.read_bytes()
    assert promoted[0].stat().st_mode & 0o777 == 0o640
    assert first["files"][0]["client_url"].endswith(
        "/r_12ab34cd/weekly-report.xlsx"
    )
    assert first["files"][0]["media_ref"] == (
        f"/media/tgg/hermes/{promoted[0].name}"
    )
    assert "r_12ab34cd" in promoted[0].name
    assert hashlib.sha256(workbook.read_bytes()).hexdigest() in promoted[0].name

    second, error = sandbox._harvest(
        work, "summary", "", "success", sandbox.DEFAULTS, **kwargs
    )
    assert error == ""
    assert second["files"][0]["media_ref"] == first["files"][0]["media_ref"]
    assert list(retained.glob("*.xlsx")) == promoted
    assert not list(retained.glob("*.tmp"))


@pytest.mark.parametrize(
    ("filename", "contents"),
    [
        ("weekly report.xlsx", b"PK\x03\x04not-promoted"),
        ("weekly-report.xlsx", b"not-a-zip"),
        ("weekly-report.csv", b"PK\x03\x04not-an-xlsx"),
    ],
)
def test_file_listing_does_not_promote_invalid_workbooks(
    tmp_path, filename, contents
):
    work = tmp_path / "work"
    retained = tmp_path / "retained"
    work.mkdir()
    (work / filename).write_bytes(contents)

    payload, _ = sandbox._harvest(
        work,
        "",
        "",
        "success",
        sandbox.DEFAULTS,
        run_id="r_12ab34cd",
        media_retention={
            "root": str(retained),
            "media_ref_prefix": "/media/tgg/hermes",
        },
    )

    assert all("media_ref" not in item for item in payload["files"])
    assert not retained.exists()


def test_sqlite_copy_larger_than_scratch_fails_clearly(tmp_path):
    source = tmp_path / "large.db"
    connection = sqlite3.connect(source)
    connection.execute("create table payload (value blob)")
    connection.execute("insert into payload values (zeroblob(2 * 1024 * 1024))")
    connection.commit()
    connection.close()

    mounts, error = sandbox._resolve_datasets(
        ["cases"],
        {
            "datasets": {"cases": {"type": "sqlite", "path": str(source)}},
            "limits": {
                "scratch_mb": 1,
                "file_size_mb": 8,
                "max_snapshot_mb": 8,
            },
        },
        tmp_path / "inputs",
    )

    assert mounts == {}
    assert "dataset 'cases'" in error
    assert "scratch_mb (1MB)" in error


def test_file_listing_omits_client_url_when_base_is_unset(tmp_path):
    work = tmp_path / "work"
    work.mkdir()
    (work / "weekly-report.xlsx").write_bytes(b"workbook")

    payload, _ = sandbox._harvest(
        work,
        "",
        "",
        "success",
        sandbox.DEFAULTS,
        run_id="r_12ab34cd",
    )

    assert all("client_url" not in item for item in payload["files"])


@pytest.mark.parametrize(
    ("run_id", "filename"),
    [
        ("r_12ab34cd", "weekly report.xlsx"),
        ("r_nothex00", "weekly-report.xlsx"),
    ],
)
def test_file_listing_omits_client_url_for_invalid_contract_values(
    tmp_path, run_id, filename
):
    work = tmp_path / "work"
    work.mkdir()
    (work / filename).write_bytes(b"workbook")

    payload, _ = sandbox._harvest(
        work,
        "",
        "",
        "success",
        sandbox.DEFAULTS,
        run_id=run_id,
        artifact_url_base="https://portal.example/artifacts",
    )

    assert all("client_url" not in item for item in payload["files"])


def test_result_and_file_metadata_are_recursively_sanitized(tmp_path, monkeypatch):
    work = tmp_path / "work"
    work.mkdir()
    (work / "result.json").write_text(
        json.dumps(
            {
                "\u001b[31msecret-value\u001b[0m": [
                    "\u001b[31msecret-value\u001b[0m"
                ]
            }
        ),
        encoding="utf-8",
    )
    (work / "artifact.txt").write_text("ok", encoding="utf-8")
    monkeypatch.setattr(
        sandbox,
        "_clean",
        lambda text: text.replace("\u001b[31m", "").replace("\u001b[0m", "").replace(
            "secret-value", "[REDACTED]"
        ),
    )
    payload, error = sandbox._harvest(
        work, "summary", "", "success", sandbox.DEFAULTS
    )
    assert error == ""
    assert payload["result"] == {"[REDACTED]": ["[REDACTED]"]}
    assert "\u001b" not in json.dumps(payload)


def test_probe_disabled_and_missing_binary(monkeypatch):
    monkeypatch.setattr(sandbox, "_load_config", lambda: {"enabled": False})
    assert sandbox._probe(force=True) == (
        False,
        "python_sandbox.enabled is false or missing",
    )
    monkeypatch.setattr(sandbox, "_load_config", lambda: {"enabled": True})
    monkeypatch.setattr(sandbox.shutil, "which", lambda _: None)
    assert sandbox._probe(force=True) == (
        False,
        "required sandbox executable(s) missing: unshare, mount, findmnt, pivot_root",
    )


def test_probe_failure_is_actionable_and_handler_fails_closed(monkeypatch):
    monkeypatch.setattr(sandbox, "_load_config", lambda: {"enabled": True})
    monkeypatch.setattr(sandbox.shutil, "which", lambda _: "/usr/bin/unshare")
    monkeypatch.setattr(sandbox.platform, "system", lambda: "Linux")
    monkeypatch.setattr(
        sandbox.subprocess,
        "run",
        lambda *args, **kwargs: Mock(returncode=1, stderr="permission denied"),
    )
    assert sandbox.check_sandbox_available() is False
    response = json.loads(sandbox._handle_python_sandbox({"code": "print(1)"}))
    assert response["status"] == "unavailable"
    assert "permission denied" in response["error"]


def test_cpu_limit_wrapper_status_is_recognized():
    assert sandbox._cpu_limit_exhausted(
        1, "unshare: sigprocmask unblock failed: Invalid argument"
    )
    assert sandbox._cpu_limit_exhausted(128 + signal.SIGXCPU, "")
    assert not sandbox._cpu_limit_exhausted(1, "ordinary script failure")


def _can_run_jail() -> bool:
    if platform.system() != "Linux":
        return False
    config = sandbox._load_config
    try:
        sandbox._load_config = lambda: {"enabled": True}
        return sandbox._probe(force=True)[0]
    finally:
        sandbox._load_config = config


@pytest.mark.skipif(not _can_run_jail(), reason="kernel user namespaces unavailable")
@pytest.mark.sandbox_e2e
def test_jailed_batch_reconciliation_e2e(tmp_path, monkeypatch):
    source, csv = tmp_path / "records.db", tmp_path / "input.csv"
    _db(source)
    _csv(csv)
    config = _config(source, csv, wall_seconds=20, max_wall_seconds=30)
    monkeypatch.setattr(sandbox, "_load_config", lambda: config)
    monkeypatch.setattr(sandbox, "get_hermes_home", lambda: tmp_path / "home")
    code = r'''
import csv, json, os, sqlite3
inputs = json.loads(os.environ["SANDBOX_INPUTS"])
assert inputs["records"] == "/work/records.db"
assert inputs["sheet"] == "/inputs/sheet"
database = sqlite3.connect(inputs["records"])
database.execute("create table writable_probe (value text)")
database.execute("insert into writable_probe values ('ok')")
database.commit()
assert database.execute("select value from writable_probe").fetchone() == ("ok",)
actual = dict(database.execute(
    "select record_no, state from records"
))
rows = list(csv.DictReader(open(inputs["sheet"], encoding="utf-8")))
missing = [r["record_no"] for r in rows if r["record_no"] not in actual]
mismatched = [r["record_no"] for r in rows
              if r["record_no"] in actual and actual[r["record_no"]] != r["state"]]
with open("/work/details.csv", "w") as f:
    f.write("kind,record_no\n")
    for value in missing: f.write("missing," + value + "\n")
    for value in mismatched: f.write("mismatch," + value + "\n")
result = {"rows": len(rows), "missing": len(missing), "mismatched": len(mismatched)}
open(os.environ["RESULT_PATH"], "w").write(json.dumps(result))
print("600 compared; 50 missing; 20 mismatched")
'''
    response = json.loads(
        sandbox.python_sandbox(code, ["records", "sheet"], timeout_seconds=20)
    )
    assert response["status"] == "success", response
    assert response["result"] == {"rows": 600, "missing": 50, "mismatched": 20}
    assert "50 missing" in response["stdout"]
    assert any(item["path"] == "work/details.csv" for item in response["files"])


@pytest.mark.skipif(not _can_run_jail(), reason="kernel user namespaces unavailable")
@pytest.mark.sandbox_e2e
def test_jailed_open_cases_xlsx_acceptance(tmp_path, monkeypatch):
    """Produce a real workbook from a copied tenant-shaped database."""
    source = tmp_path / "tenant-source.db"
    connection = sqlite3.connect(source)
    connection.execute(
        "create table cases ("
        "job_no text primary key, zone text not null, state text not null, "
        "priority text, updated_at integer not null)"
    )
    connection.executemany(
        "insert into cases values (?, ?, ?, ?, ?)",
        [
            (
                f"JOB-{index:04d}",
                f"ZONE-{index % 4 + 1}",
                "open" if index % 3 else "closed",
                "urgent" if index % 10 == 1 else "normal",
                1_800_000_000 + index,
            )
            for index in range(120)
        ],
    )
    connection.commit()
    connection.close()

    tenant_copy = tmp_path / "tenant-copy.db"
    shutil.copy2(source, tenant_copy)
    config = {
        "enabled": True,
        "datasets": {
            "cases": {
                "type": "sqlite",
                "path": str(tenant_copy),
                "description": "Copied tenant case database.",
            },
        },
        "limits": {"wall_seconds": 30, "max_wall_seconds": 30},
    }
    home = tmp_path / "home"
    monkeypatch.setattr(sandbox, "_load_config", lambda: config)
    monkeypatch.setattr(sandbox, "get_hermes_home", lambda: home)
    code = r'''
import json, os, sqlite3
from openpyxl import Workbook

inputs = json.loads(os.environ["SANDBOX_INPUTS"])
rows = sqlite3.connect(inputs["cases"]).execute(
    "select job_no, zone, state, priority, updated_at "
    "from cases where state = 'open' order by job_no"
).fetchall()
workbook = Workbook()
sheet = workbook.active
sheet.title = "Open Cases"
sheet.append(["Job No", "Zone", "State", "Priority", "Updated At"])
for row in rows:
    sheet.append(row)
sheet.freeze_panes = "A2"
sheet.auto_filter.ref = sheet.dimensions
sheet.column_dimensions["A"].width = 18
sheet.column_dimensions["B"].width = 14
sheet.column_dimensions["C"].width = 12
sheet.column_dimensions["D"].width = 12
sheet.column_dimensions["E"].width = 18
output = "/work/open-cases.xlsx"
workbook.save(output)
result = {"open_cases": len(rows), "workbook": "open-cases.xlsx"}
open(os.environ["RESULT_PATH"], "w").write(json.dumps(result))
print(f"{len(rows)} open cases exported to open-cases.xlsx")
'''
    response = json.loads(
        sandbox.python_sandbox(code, ["cases"], timeout_seconds=30)
    )
    assert response["status"] == "success", response
    assert response["result"] == {
        "open_cases": 80,
        "workbook": "open-cases.xlsx",
    }
    assert any(
        item["path"] == "work/open-cases.xlsx" and item["bytes"] > 0
        for item in response["files"]
    )

    output = home / "sandbox_runs" / response["run_id"] / "work" / "open-cases.xlsx"
    from openpyxl import load_workbook

    workbook = load_workbook(output, read_only=True)
    sheet = workbook["Open Cases"]
    assert sheet.max_row == 81
    assert tuple(cell.value for cell in next(sheet.iter_rows(max_row=1))) == (
        "Job No",
        "Zone",
        "State",
        "Priority",
        "Updated At",
    )
    assert {row[2] for row in sheet.iter_rows(min_row=2, values_only=True)} == {
        "open"
    }
    workbook.close()

    artifact_dir = os.environ.get("PYTHON_SANDBOX_ACCEPTANCE_DIR")
    if artifact_dir:
        destination = Path(artifact_dir)
        destination.mkdir(parents=True, exist_ok=True)
        shutil.copy2(output, destination / "open-cases.xlsx")


@pytest.mark.skipif(not _can_run_jail(), reason="kernel user namespaces unavailable")
@pytest.mark.sandbox_e2e
def test_jailed_network_and_write_escape_are_blocked(tmp_path, monkeypatch):
    source, csv = tmp_path / "records.db", tmp_path / "input.csv"
    _db(source)
    _csv(csv)
    media = tmp_path / "media"
    media.mkdir()
    (media / "sample.txt").write_text("read-only", encoding="utf-8")
    config = _config(source, csv, wall_seconds=20)
    config["datasets"]["media"] = {
        "type": "path",
        "path": str(media),
        "description": "Generic directory dataset.",
    }
    home = tmp_path / "home"
    monkeypatch.setattr(sandbox, "_load_config", lambda: config)
    monkeypatch.setattr(sandbox, "get_hermes_home", lambda: home)
    code = r'''
import ctypes, errno, json, os, socket, urllib.request
inputs = json.loads(os.environ["SANDBOX_INPUTS"])
blocked = {}
for label, action in {
    "network": lambda: socket.create_connection(("1.1.1.1", 53), 1),
    "http": lambda: urllib.request.urlopen("http://1.1.1.1", timeout=1),
    "input_write": lambda: open(inputs["sheet"], "w"),
    "directory_write": lambda: open(inputs["media"] + "/new.txt", "w"),
    "etc_write": lambda: open("/etc/x", "w"),
    "home_is_scratch": lambda: (
        os.path.expanduser("~") == "/work"
        and open(os.path.expanduser("~/x"), "w").close() is None
    ),
    "pid1_environ": lambda: open("/proc/1/environ", "rb").read(),
    "venv_write": lambda: open("/venv/probe.txt", "w"),
}.items():
    try: action(); blocked[label] = False
    except Exception: blocked[label] = True
libc = ctypes.CDLL(None, use_errno=True)
libc.mount.argtypes = [ctypes.c_char_p, ctypes.c_char_p, ctypes.c_char_p,
                       ctypes.c_ulong, ctypes.c_void_p]
for label, target in {
    "venv_remount": b"/venv",
    "directory_remount": inputs["media"].encode(),
}.items():
    ctypes.set_errno(0)
    rc = libc.mount(None, target, None, 32 | 4096, None)
    blocked[label] = rc == -1 and ctypes.get_errno() == errno.EPERM
ctypes.set_errno(0)
rc = libc.ptrace(16, 1, None, None)  # PTRACE_ATTACH namespace PID 1
blocked["ptrace_pid1"] = rc == -1 and ctypes.get_errno() == errno.EPERM
open("/work/ok.txt", "w").write("ok")
os.symlink("/etc/passwd", "/work/host-pointer")
os.mkfifo("/work/host-fifo")
open(os.environ["RESULT_PATH"], "w").write(json.dumps(blocked))
'''
    response = json.loads(
        sandbox.python_sandbox(code, ["sheet", "media"], timeout_seconds=20)
    )
    assert response["status"] == "success", response
    assert response["result"] == {
        "network": True,
        "http": True,
        "input_write": True,
        "directory_write": True,
        "etc_write": True,
        "home_is_scratch": False,
        "pid1_environ": True,
        "ptrace_pid1": True,
        "venv_write": True,
        "venv_remount": True,
        "directory_remount": True,
    }
    assert any(item["path"] == "work/ok.txt" for item in response["files"])
    assert all(item["path"] != "work/host-pointer" for item in response["files"])
    assert all(item["path"] != "work/host-fifo" for item in response["files"])
    run_work = home / "sandbox_runs" / response["run_id"] / "work"
    assert not os.path.lexists(run_work / "host-pointer")
    assert not os.path.lexists(run_work / "host-fifo")


@pytest.mark.skipif(not _can_run_jail(), reason="kernel user namespaces unavailable")
@pytest.mark.sandbox_e2e
def test_jailed_host_paths_and_home_canary_are_absent(tmp_path, monkeypatch):
    source, csv = tmp_path / "records.db", tmp_path / "input.csv"
    _db(source)
    _csv(csv)
    home = tmp_path / "fake-home"
    home.mkdir()
    canary = home / "must-not-be-visible"
    canary.write_text("canary", encoding="utf-8")
    config = _config(source, csv, wall_seconds=20)
    monkeypatch.setattr(sandbox, "_load_config", lambda: config)
    monkeypatch.setattr(sandbox, "get_hermes_home", lambda: home)
    code = f'''
import json, os
checks = {{
    "home_canary_absent": not os.path.exists({str(canary)!r}),
    "live_database_absent": not os.path.exists({str(source)!r}),
    "snapshot_present": os.path.exists(json.loads(os.environ["SANDBOX_INPUTS"])["records"]),
}}
open(os.environ["RESULT_PATH"], "w").write(json.dumps(checks))
'''
    response = json.loads(sandbox.python_sandbox(code, ["records"], timeout_seconds=20))
    assert response["status"] == "success", response
    assert response["result"] == {
        "home_canary_absent": True,
        "live_database_absent": True,
        "snapshot_present": True,
    }


@pytest.mark.skipif(not _can_run_jail(), reason="kernel user namespaces unavailable")
@pytest.mark.sandbox_e2e
def test_jailed_timeout_is_bounded_and_leaves_no_process(tmp_path, monkeypatch):
    source, csv = tmp_path / "records.db", tmp_path / "input.csv"
    _db(source)
    _csv(csv)
    home = tmp_path / "home"
    config = _config(source, csv, wall_seconds=5, max_wall_seconds=5)
    monkeypatch.setattr(sandbox, "_load_config", lambda: config)
    monkeypatch.setattr(sandbox, "get_hermes_home", lambda: home)
    started = time.monotonic()
    response = json.loads(
        sandbox.python_sandbox("while True: pass", timeout_seconds=5)
    )
    elapsed = time.monotonic() - started
    assert response["status"] == "timeout", response
    assert elapsed <= 12
    run_path = str(home / "sandbox_runs" / response["run_id"])
    processes = subprocess.run(
        ["ps", "-eo", "args="], capture_output=True, text=True, check=True
    ).stdout
    assert run_path not in processes


@pytest.mark.skipif(not _can_run_jail(), reason="kernel user namespaces unavailable")
@pytest.mark.sandbox_e2e
def test_jailed_oom_has_distinct_status_and_guidance(tmp_path, monkeypatch):
    source, csv = tmp_path / "records.db", tmp_path / "input.csv"
    _db(source)
    _csv(csv)
    config = _config(
        source,
        csv,
        wall_seconds=20,
        max_wall_seconds=20,
        memory_mb=96,
    )
    monkeypatch.setattr(sandbox, "_load_config", lambda: config)
    monkeypatch.setattr(sandbox, "get_hermes_home", lambda: tmp_path / "home")
    response = json.loads(
        sandbox.python_sandbox(
            "x = bytearray(512 * 1024 * 1024)", timeout_seconds=20
        )
    )
    assert response["status"] == "oom", response
    assert "stream/chunk" in response["error"]


@pytest.mark.skipif(not _can_run_jail(), reason="kernel user namespaces unavailable")
@pytest.mark.sandbox_e2e
def test_jailed_cpu_limit_has_distinct_guidance(tmp_path, monkeypatch):
    source, csv = tmp_path / "records.db", tmp_path / "input.csv"
    _db(source)
    _csv(csv)
    config = _config(
        source,
        csv,
        wall_seconds=10,
        max_wall_seconds=10,
        cpu_seconds=1,
    )
    monkeypatch.setattr(sandbox, "_load_config", lambda: config)
    monkeypatch.setattr(sandbox, "get_hermes_home", lambda: tmp_path / "home")
    response = json.loads(
        sandbox.python_sandbox("while True: pass", timeout_seconds=10)
    )
    assert response["status"] == "error", response
    assert "CPU limit (1s) exhausted" in response["stderr"]


@pytest.mark.skipif(not _can_run_jail(), reason="kernel user namespaces unavailable")
@pytest.mark.sandbox_e2e
def test_jailed_process_count_is_bounded(tmp_path, monkeypatch):
    source, csv = tmp_path / "records.db", tmp_path / "input.csv"
    _db(source)
    _csv(csv)
    config = _config(
        source,
        csv,
        wall_seconds=20,
        max_wall_seconds=20,
        max_processes=16,
    )
    monkeypatch.setattr(sandbox, "_load_config", lambda: config)
    monkeypatch.setattr(sandbox, "get_hermes_home", lambda: tmp_path / "home")
    code = r'''
import errno, json, os, subprocess
children = []
errors = []
for _ in range(40):
    try:
        children.append(subprocess.Popen(
            ["/venv/bin/python", "-c", "import time; time.sleep(5)"],
        ))
    except OSError as exc:
        errors.append(exc.errno)
for child in children:
    child.terminate()
for child in children:
    child.wait()
open(os.environ["RESULT_PATH"], "w").write(json.dumps(
    {"spawned": len(children), "errors": errors}
))
'''
    response = json.loads(sandbox.python_sandbox(code, timeout_seconds=20))
    assert response["status"] == "success", response
    assert 1 <= response["result"]["spawned"] <= 15
    assert errno.EAGAIN in response["result"]["errors"]


@pytest.mark.skipif(not _can_run_jail(), reason="kernel user namespaces unavailable")
@pytest.mark.sandbox_e2e
def test_jailed_scratch_total_is_bounded(tmp_path, monkeypatch):
    source, csv = tmp_path / "records.db", tmp_path / "input.csv"
    _db(source)
    _csv(csv)
    home = tmp_path / "home"
    config = _config(
        source,
        csv,
        wall_seconds=20,
        max_wall_seconds=20,
        file_size_mb=4,
        scratch_mb=4,
    )
    config["workspace"] = "session"
    monkeypatch.setattr(sandbox, "_load_config", lambda: config)
    monkeypatch.setattr(sandbox, "get_hermes_home", lambda: home)
    code = r'''
import errno, json, os
written = []
error = None
for index in range(8):
    path = f"/work/chunk-{index}.bin"
    try:
        with open(path, "wb") as handle:
            handle.write(b"x" * (1024 * 1024))
        written.append(path)
    except OSError as exc:
        error = exc.errno
        try:
            os.unlink(path)
        except FileNotFoundError:
            pass
        break
if written:
    os.unlink(written.pop())
open(os.environ["RESULT_PATH"], "w").write(json.dumps(
    {"written": len(written), "error": error}
))
'''
    response = json.loads(
        sandbox.python_sandbox(
            code,
            timeout_seconds=20,
            session_id="scratch-budget-chat",
        )
    )
    assert response["status"] == "success", response
    assert response["result"]["error"] == errno.ENOSPC
    run_work = home / "sandbox_runs" / response["run_id"] / "work"
    assert sum(path.stat().st_size for path in run_work.rglob("*") if path.is_file()) <= (
        4 * 1024 * 1024
    )
    retained_work = (
        home
        / "sandbox_workspaces"
        / sandbox._workspace_key("scratch-budget-chat")
        / "work"
    )
    assert sum(
        path.stat().st_size for path in retained_work.rglob("*") if path.is_file()
    ) <= (4 * 1024 * 1024)


def test_prune_removes_expired_and_excess_runs(tmp_path):
    root = tmp_path / "runs"
    root.mkdir()
    for index in range(3):
        path = root / f"r_{index}"
        path.mkdir()
        os.utime(path, (time.time() - index * 10, time.time() - index * 10))
    sandbox._prune_runs(root, {"artifact_ttl_days": 0, "max_runs_kept": 2})
    assert len(list(root.iterdir())) == 2


def test_prune_workspaces_removes_expired_and_excess_sessions(tmp_path):
    root = tmp_path / "workspaces"
    root.mkdir()
    now = time.time()
    for index in range(4):
        path = root / f"s_{index}"
        path.mkdir()
        age = 3 * 86400 if index == 3 else index * 10
        os.utime(path, (now - age, now - age))
    sandbox._prune_workspaces(
        root,
        {"artifact_ttl_days": 1, "max_session_workspaces": 10},
    )
    assert {path.name for path in root.iterdir()} == {"s_0", "s_1", "s_2"}
    sandbox._prune_workspaces(
        root,
        {"artifact_ttl_days": 0, "max_session_workspaces": 2},
    )
    assert {path.name for path in root.iterdir()} == {"s_0", "s_1"}


@pytest.mark.skipif(not _can_run_jail(), reason="kernel user namespaces unavailable")
@pytest.mark.sandbox_e2e
def test_session_workspace_persists_and_isolates_across_runs(tmp_path, monkeypatch):
    source, csv = tmp_path / "records.db", tmp_path / "input.csv"
    _db(source)
    _csv(csv)
    config = _config(
        source,
        csv,
        wall_seconds=20,
        max_wall_seconds=20,
        scratch_mb=4,
    )
    config.update(
        {
            "workspace": "session",
            "artifact_ttl_days": 7,
            "max_session_workspaces": 4,
        }
    )
    home = tmp_path / "home"
    monkeypatch.setattr(sandbox, "_load_config", lambda: config)
    monkeypatch.setattr(sandbox, "get_hermes_home", lambda: home)

    first = json.loads(
        sandbox.python_sandbox(
            'open("/work/state.txt", "w").write("alpha")',
            timeout_seconds=20,
            session_id="chat-a",
        )
    )
    assert first["status"] == "success", first

    second = json.loads(
        sandbox.python_sandbox(
            'import json,os; open(os.environ["RESULT_PATH"], "w").write('
            'json.dumps({"state": open("/work/state.txt").read()}))',
            timeout_seconds=20,
            session_id="chat-a",
        )
    )
    assert second["status"] == "success", second
    assert second["result"] == {"state": "alpha"}

    no_stale_result = json.loads(
        sandbox.python_sandbox(
            'assert open("/work/state.txt").read() == "alpha"',
            timeout_seconds=20,
            session_id="chat-a",
        )
    )
    assert no_stale_result["status"] == "success", no_stale_result
    assert no_stale_result["result"] is None

    isolated = json.loads(
        sandbox.python_sandbox(
            'import json,os; open(os.environ["RESULT_PATH"], "w").write('
            'json.dumps({"exists": os.path.exists("/work/state.txt")}))',
            timeout_seconds=20,
            session_id="chat-b",
        )
    )
    assert isolated["status"] == "success", isolated
    assert isolated["result"] == {"exists": False}
    workspace_root = home / "sandbox_workspaces"
    assert len(list(workspace_root.iterdir())) == 2


@pytest.mark.skipif(not _can_run_jail(), reason="kernel user namespaces unavailable")
@pytest.mark.sandbox_e2e
def test_invalid_dataset_does_not_create_session_workspace(tmp_path, monkeypatch):
    source, csv = tmp_path / "records.db", tmp_path / "input.csv"
    _db(source)
    _csv(csv)
    config = _config(source, csv)
    config["workspace"] = "session"
    home = tmp_path / "home"
    monkeypatch.setattr(sandbox, "_load_config", lambda: config)
    monkeypatch.setattr(sandbox, "get_hermes_home", lambda: home)
    response = json.loads(
        sandbox.python_sandbox(
            "print(1)",
            datasets=["missing"],
            session_id="bad-dataset-chat",
        )
    )
    assert response["status"] == "dataset_unknown"
    assert not (home / "sandbox_workspaces").exists()
