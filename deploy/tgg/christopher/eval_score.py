#!/usr/bin/env python3
"""Score Christopher model arms against pinned TGG parallel artifacts.

This is the TGG payload adapter for ``gateway.eval_instrument``.  It reuses the
field and status semantics from the June decision-surface builder while keeping
client data outside git: inputs are paths, and the emitted score contains only
counts plus hashed case identities.
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import re
import sqlite3
import urllib.parse
from pathlib import Path
from typing import Any, Iterable, Mapping

import openpyxl

from gateway.eval_instrument import (
    SCORE_SCHEMA,
    evaluate_adaptive_trace,
    load_eval_config,
    pin_eval_config,
)


OPERATIONAL_TABS = {"Punggol": "PG", "Hougang": "HG", "Bedok": "HG"}
JOB_PATTERN = re.compile(
    r"(?:AM|SK|PG|HG|BI)\s*/?\s*JOB\s*/?\s*\d{4}\s*/?\s*\d{3,5}", re.I
)
DATE_PATTERN = re.compile(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})")


def norm_job(value: Any) -> str:
    """Canonical TGG comparison key from extract_master.py/build_decision_surface.py."""
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value)).strip()


def _parse_date(value: Any) -> dt.date | None:
    if value in (None, ""):
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = _clean(value)
    for fmt in (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%d-%m-%Y",
        "%d-%m-%y",
        "%Y-%m-%d %H:%M:%S",
    ):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    match = DATE_PATTERN.search(text)
    if not match:
        return None
    day, month, year = map(int, match.groups())
    if year < 100:
        year += 2000
    try:
        return dt.date(year, month, day)
    except ValueError:
        return None


def _find_header_col(ws, names: Iterable[str]) -> int | None:
    wanted = [re.sub(r"\s+", " ", name.strip().lower()) for name in names]
    for row in range(1, min(ws.max_row, 8) + 1):
        for column in range(1, ws.max_column + 1):
            value = re.sub(r"\s+", " ", _clean(ws.cell(row, column).value).lower())
            if any(value == name or value.startswith(name) for name in wanted):
                return column
    return None


def _is_completed(*values: Any) -> bool:
    text = " ".join(_clean(value).lower() for value in values)
    return any(token in text for token in ("completed", "closed", "owner settled"))


def load_master(path: Path) -> dict[str, dict[str, Any]]:
    """Load the PG/HG master tabs using the canonical tab-to-chat mapping."""
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    rows: dict[str, dict[str, Any]] = {}
    try:
        for tab, zone in OPERATIONAL_TABS.items():
            if tab not in workbook.sheetnames:
                continue
            ws = workbook[tab]
            job_col = _find_header_col(ws, ("Job No.", "Job No")) or 3
            status_col = _find_header_col(ws, ("Status",)) or 26
            completion_col = _find_header_col(ws, ("Completion Date",)) or 23
            for row in range(1, ws.max_row + 1):
                job = _clean(ws.cell(row, job_col).value).strip("'\"")
                if not JOB_PATTERN.search(job):
                    continue
                key = norm_job(job)
                completion = _parse_date(ws.cell(row, completion_col).value)
                completed = bool(completion) or _is_completed(
                    ws.cell(row, status_col).value
                )
                prior = rows.get(key)
                candidate = {
                    "zone": zone,
                    "completed": completed,
                    "completion_date": completion,
                }
                if prior is None or (completion or dt.date.min) >= (
                    prior.get("completion_date") or dt.date.min
                ):
                    rows[key] = candidate
    finally:
        workbook.close()
    return rows


def _closure_events(value: Any) -> list[tuple[dt.date | None, str]]:
    events: list[tuple[dt.date | None, str]] = []
    for line in re.split(r"[\n\r]+", _clean(value)):
        if not line:
            continue
        spans = list(
            re.finditer(
                r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s*[-–]\s*(.*?)(?=\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\s*[-–]|$)",
                line,
                flags=re.I,
            )
        )
        fragments = [(match.group(1), match.group(2)) for match in spans] or [
            (None, line)
        ]
        for date_text, fragment in fragments:
            lowered = fragment.lower()
            if "modification" in lowered or "wc mod" in lowered:
                kind = "wc_mod"
            elif "closure" in lowered or "close" in lowered:
                kind = "closure"
            else:
                continue
            events.append((_parse_date(date_text or fragment), kind))
    return events


def load_closure(path: Path) -> dict[str, dict[str, Any]]:
    """Load latest-dated closure state; WC modification remains open."""
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    rows: dict[str, dict[str, Any]] = {}
    try:
        for ws in workbook.worksheets:
            header_row = None
            for row in range(1, min(ws.max_row, 8) + 1):
                values = [
                    _clean(ws.cell(row, column).value).lower()
                    for column in range(1, min(ws.max_column, 10) + 1)
                ]
                if "cost" in values and any(value in {"close", "closure"} for value in values):
                    header_row = row
                    break
            if header_row is None:
                continue
            headers = {
                _clean(ws.cell(header_row, column).value).lower(): column
                for column in range(1, ws.max_column + 1)
            }
            job_col = headers.get("cost") or 2
            close_col = headers.get("close") or headers.get("closure") or 4
            for row in range(header_row + 1, ws.max_row + 1):
                job = _clean(ws.cell(row, job_col).value).strip("'\"")
                if not JOB_PATTERN.search(job):
                    continue
                events = _closure_events(ws.cell(row, close_col).value)
                if not events:
                    continue
                dated = sorted(
                    ((date, kind) for date, kind in events if date is not None),
                    key=lambda item: item[0],
                )
                if dated:
                    date, kind = dated[-1]
                else:
                    date = None
                    kind = "closure" if any(kind == "closure" for _, kind in events) else "wc_mod"
                key = norm_job(job)
                prior = rows.get(key)
                if prior is None or (date or dt.date.min) >= (
                    prior.get("date") or dt.date.min
                ):
                    rows[key] = {
                        "completed": kind == "closure",
                        "date": date,
                        "kind": kind,
                    }
    finally:
        workbook.close()
    return rows


def _read_capture_job_keys(paths: Iterable[Path]) -> set[str]:
    keys: set[str] = set()
    for path in paths:
        with path.open(encoding="utf-8") as handle:
            for line in handle:
                if not line.strip():
                    continue
                record = json.loads(line)
                normalized = record.get("normalized") if isinstance(record, Mapping) else None
                body = normalized.get("body") if isinstance(normalized, Mapping) else ""
                for match in JOB_PATTERN.finditer(str(body or "")):
                    keys.add(norm_job(match.group(0)))
    return keys


def _readonly(path: Path) -> sqlite3.Connection:
    connection = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
    connection.row_factory = sqlite3.Row
    return connection


def _case_rows(path: Path) -> dict[str, dict[str, Any]]:
    with _readonly(path) as connection:
        return {
            norm_job(row["job_no"] or row["normalized_job_no"]): dict(row)
            for row in connection.execute(
                "SELECT id, job_no, normalized_job_no, state, job_status, linkfm_status FROM cases"
            )
            if norm_job(row["job_no"] or row["normalized_job_no"])
        }


def _touched_case_keys(target_db: Path, run_id: str) -> set[str]:
    touched: set[str] = set()
    with _readonly(target_db) as connection:
        for row in connection.execute(
            """
            SELECT c.job_no, c.normalized_job_no
            FROM case_observations o JOIN cases c ON c.id = o.case_id
            WHERE json_extract(o.fields, '$.replay_run_id') = ?
            """,
            (run_id,),
        ):
            touched.add(norm_job(row["job_no"] or row["normalized_job_no"]))
        for row in connection.execute(
            "SELECT path FROM replay_target_tool_writes WHERE run_id = ? AND status < 400",
            (run_id,),
        ):
            match = re.search(r"/cases/([^/?]+)(?:/|$)", str(row["path"] or ""))
            if match:
                touched.add(norm_job(urllib.parse.unquote(match.group(1))))
    return {key for key in touched if key}


def _turn_metrics(session_db: Path, run_id: str, agent_id: str) -> dict[str, Any]:
    with _readonly(session_db) as connection:
        rows = list(
            connection.execute(
                """
                SELECT cost_usd, latency_ms, turn_status
                FROM pa_turns
                WHERE replay_run_id = ? AND agent_id = ?
                """,
                (run_id, agent_id),
            )
        )
    latencies = sorted(
        int(row["latency_ms"])
        for row in rows
        if row["latency_ms"] is not None
    )
    if latencies:
        index = max(0, min(len(latencies) - 1, int((len(latencies) - 1) * 0.95)))
        p95 = latencies[index]
    else:
        p95 = None
    return {
        "total_cost_usd": round(
            sum(float(row["cost_usd"] or 0) for row in rows), 6
        ),
        "p95_turn_latency_ms": p95,
        "failed_turns": sum(row["turn_status"] == "failed" for row in rows),
        "turns": len(rows),
    }


def _completed(case: Mapping[str, Any]) -> bool:
    return _is_completed(case.get("state"), case.get("job_status"), case.get("linkfm_status"))


def _case_token(key: str) -> str:
    return hashlib.sha256(key.encode("utf-8")).hexdigest()[:16]


def build_score(args: argparse.Namespace) -> dict[str, Any]:
    config = load_eval_config(args.config)
    pins = pin_eval_config(config)
    run_manifest = json.loads(Path(args.run_manifest).read_text(encoding="utf-8"))
    run_id = str(run_manifest.get("run_id") or "")
    if not run_id:
        raise ValueError("run manifest is missing run_id")
    source_paths = [
        Path(source["path"]).expanduser().resolve()
        for source in config.data["corpus"]["sources"]
    ]
    corpus_job_keys = _read_capture_job_keys(source_paths)
    master = load_master(Path(args.master).expanduser().resolve())
    closure = load_closure(Path(args.closure).expanduser().resolve())
    tracker_keys = set(master) | set(closure)
    baseline = _case_rows(Path(args.baseline_db).expanduser().resolve())
    target = _case_rows(Path(args.target_db).expanduser().resolve())
    new_keys = set(target) - set(baseline)
    touched = _touched_case_keys(Path(args.target_db).expanduser().resolve(), run_id)
    matched = touched & set(baseline)

    tracker_miss = {key for key in new_keys if key not in tracker_keys}
    phantom = {key for key in tracker_miss if key not in corpus_job_keys}
    wrong_completion = set()
    future_resolution_excluded = set()
    for key in touched | new_keys:
        case = target.get(key)
        if not case or not _completed(case) or key not in tracker_keys:
            continue
        closure_row = closure.get(key)
        master_row = master.get(key)
        tracker_completed = (
            closure_row["completed"]
            if closure_row is not None
            else bool(master_row and master_row["completed"])
        )
        if not tracker_completed:
            wrong_completion.add(key)
            continue
        resolution_date = (
            closure_row.get("date")
            if closure_row is not None
            else master_row.get("completion_date") if master_row else None
        )
        if resolution_date and resolution_date >= dt.date(2026, 5, 24):
            future_resolution_excluded.add(key)

    trace = evaluate_adaptive_trace(
        config,
        args.arm,
        run_manifest_path=args.run_manifest,
        session_db_path=args.session_db,
    )
    paired_check = next(
        (check for check in trace["checks"] if check["name"] == "paired-probes"),
        {"actual": {"probes": []}},
    )
    twin_rows = [
        {
            "probe_id": probe["id"],
            "passed": bool(probe["ok"]),
            "outcome": "different_paths" if probe["ok"] else "same_or_missing_path",
        }
        for probe in paired_check.get("actual", {}).get("probes", [])
    ]
    turn_metrics = _turn_metrics(
        Path(args.session_db).expanduser().resolve(),
        run_id,
        pins["agent"]["id"],
    )
    cases = []
    for key in sorted(new_keys | matched | wrong_completion):
        outcomes = []
        if key in new_keys:
            outcomes.append("created")
        if key in matched:
            outcomes.append("matched_existing")
        if key in tracker_miss:
            outcomes.append("tracker_miss_candidate")
        if key in phantom:
            outcomes.append("phantom_candidate")
        if key in wrong_completion:
            outcomes.append("wrong_completion")
        if key in future_resolution_excluded:
            outcomes.append("future_resolution_excluded")
        cases.append(
            {
                "case_id": _case_token(key),
                "outcome": "+".join(outcomes),
                "correct": not bool(
                    key in phantom or key in wrong_completion
                ),
            }
        )
    metrics = {
        "cases_created": len(new_keys),
        "cases_matched": len(matched),
        "phantom_cases": len(phantom),
        "wrong_completions": len(wrong_completion),
        **{key: turn_metrics[key] for key in ("total_cost_usd", "p95_turn_latency_ms", "failed_turns")},
    }
    return {
        "schema": SCORE_SCHEMA,
        "arm_id": args.arm,
        "eligible": True,
        "instrument_id": config.data["instrument_id"],
        "instrument_digest": pins["digest"],
        "run_id": run_id,
        "judge_digest": (pins.get("judge") or {}).get("digest"),
        "metrics": metrics,
        "cases": cases,
        "twin_discrimination": twin_rows,
        "audit": {
            "turns": turn_metrics["turns"],
            "tracker_miss_candidates": len(tracker_miss),
            "future_resolution_excluded": len(future_resolution_excluded),
            "client_identifiers_emitted": False,
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", required=True)
    parser.add_argument("--arm", required=True)
    parser.add_argument("--run-manifest", required=True)
    parser.add_argument("--session-db", required=True)
    parser.add_argument("--baseline-db", required=True)
    parser.add_argument("--target-db", required=True)
    parser.add_argument("--master", required=True)
    parser.add_argument("--closure", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    score = build_score(args)
    output = Path(args.output).expanduser().resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        json.dumps(score, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )
    print(
        json.dumps(
            {
                "ok": True,
                "output": str(output),
                "sha256": hashlib.sha256(output.read_bytes()).hexdigest(),
                "metrics": score["metrics"],
                "twin_discrimination": score["twin_discrimination"],
            },
            ensure_ascii=False,
            indent=2,
            sort_keys=True,
        )
    )


if __name__ == "__main__":
    main()
